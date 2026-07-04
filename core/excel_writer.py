"""
excel_writer.py
---------------
Writes all tenders to Excel incrementally. Supports live saving,
auto‑recreation if the file is deleted, and clean multiline text handling.
"""

import os
from datetime import datetime
from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from core.logger import get_logger

logger = get_logger(__name__)

COLUMN_HEADERS = [
    "Sr No", "Bid Number", "Dated", "Bid End Date",
    "Contract Period", "Address", "Eligible"
]

# Characters illegal in Excel cells (except \t, \n, \r)
ILLEGAL_CHARS = set(range(0, 9)).union(range(11, 13)).union(range(14, 32)).union({127, 160})


def sanitize(value: str) -> str:
    if not isinstance(value, str):
        value = str(value)
    # Remove illegal characters
    clean = "".join(c for c in value if ord(c) not in ILLEGAL_CHARS)
    # Replace newlines with ", " for a single-line display
    return clean.replace("\n", ", ").replace("\r", ", ")


class ExcelWriter:
    def __init__(self, output_path: str):
        self.output_path = output_path
        self.fallback_path = None
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "All Tenders"
        self.current_row = 2
        self.sr_no = 0
        self._header_written = False
        self._save_failed = False

    def _write_header(self):
        fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        font = Font(bold=True, color="FFFFFF")
        align = Alignment(horizontal="center", vertical="center")
        for col, header in enumerate(COLUMN_HEADERS, 1):
            cell = self.sheet.cell(row=1, column=col, value=header)
            cell.font = font; cell.fill = fill; cell.alignment = align
        self.sheet.freeze_panes = "A2"
        self._header_written = True

    def add_tender(self, tender: Dict[str, str], save_after: bool = False):
        if not self._header_written:
            self._write_header()

        self.sr_no += 1
        row = self.current_row
        center = Alignment(horizontal="center", vertical="center")

        self.sheet.cell(row=row, column=1, value=self.sr_no).alignment = center
        self.sheet.cell(row=row, column=2, value=sanitize(tender.get("bid_number", "")))
        self.sheet.cell(row=row, column=3, value=sanitize(tender.get("dated", "")))
        self.sheet.cell(row=row, column=4, value=sanitize(tender.get("bid_end_date", "")))
        self.sheet.cell(row=row, column=5, value=sanitize(tender.get("contract_period", "")))
        self.sheet.cell(row=row, column=6, value=sanitize(tender.get("address", "")))
        self.sheet.cell(row=row, column=7, value=sanitize(tender.get("eligible", ""))).alignment = center

        self.current_row += 1
        if save_after:
            self._save()

    def _try_save(self, path: str) -> bool:
        try:
            self.workbook.save(path)
            return True
        except PermissionError:
            return False

    def _save(self):
        # Try original path; if deleted, recreate it
        if not os.path.exists(self.output_path) and self._save_failed:
            if self._try_save(self.output_path):
                logger.info("Recreated original Excel file: %s", self.output_path)
                self._save_failed = False
                self.fallback_path = None
                return

        if self._save_failed and self.fallback_path:
            if self._try_save(self.fallback_path):
                return
            else:
                logger.warning("Fallback file locked – will retry later.")
                return

        if self._try_save(self.output_path):
            return

        # Original locked → create fallback
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(self.output_path)
        self.fallback_path = f"{base}_{timestamp}{ext}"
        logger.warning("Original file locked. Switching to fallback: %s", self.fallback_path)
        self._save_failed = True
        if self._try_save(self.fallback_path):
            logger.info("Workbook saved to fallback.")
        else:
            logger.error("Fallback file also locked – data not saved!")

    def finalize(self):
        self._autosize_columns()
        self._save()
        path = self.fallback_path if self._save_failed else self.output_path
        logger.info("Final Excel saved: %s", path)

    def _autosize_columns(self, sheet=None):
        sheet = sheet or self.sheet
        for col in range(1, len(COLUMN_HEADERS) + 1):
            letter = get_column_letter(col)
            max_len = 0
            for cell in sheet[letter]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[letter].width = max_len + 4
